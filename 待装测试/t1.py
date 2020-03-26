#!/usr/bin/env python
# -*- coding: utf-8 -*-
""""""""""""""""""""""""""
# @版    本：V1.0
# @作    者：林昕明
# @联系方式：humen@189.cn
# @创建软件：PyCharm
# @项目名称：待装测试
# @文 件 名：t1.py
# @修改时间：2019/3/6 20:29
# @文件说明:而高科技开发dfe fjiezesid fieiw sl
""""""""""""""""""""""""""
import openpyxl as op
from openpyxl.styles import colors
from openpyxl.styles import Font
import datetime
wb=op.Workbook()
ws1=wb.active
ws1.title='第一页'
ft=Font(color=colors.RED)
ws2=wb.create_sheet('增加一页',0)
ws2.title='第二页'
ws1['a1']='第一页的第一个单元格'
ws2['a1']=datetime.date.today()
ws3=wb.create_sheet('第三页',0)
ws3=wb.copy_worksheet(ws1)
ws3['a1'].font=ft
wb.save('test.xls')
FileFinder(path, *loader_details)
print('fiefwifief ')


