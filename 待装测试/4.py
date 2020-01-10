import openpyxl
import openpyxl.utils
from openpyxl.styles import Font,colors,Alignment
# wb=openpyxl.Workbook()
# ws=wb.active
# wb.create_sheet('测试页',0)
# ws.title='这是修改后表格名称'
wb=openpyxl.load_workbook('电子表格测试.xlsx')
ws=wb.active #激活左边第一张表格
ws=wb['这是修改后表格名称'] #直接选择要操作的表格
# ws.append([9,9,9])
# ws.append([8,8,8])
# style=Font(name='微软雅黑', size=24, italic=True, color=colors.RED, bold=True)
# ws['a1'].font=style
# ws['a1'].alignment=Alignment(horizontal='center',vertical='center')
# ws.row_dimensions[2].height=20
# ws.column_dimensions['b'].width=25
# ws.sheet_properties.tabColor="1272BA"
# print(ws.title,wb.sheetnames)
# print(ws.max_row)
# print(ws.max_column)
# d = ws.cell(row=4, column=2, value=999)
# print(d.value)
# for i in range(1,11): 
#     for j in range(1,11): 
#         ws.cell(row=i,column=j,value='填充物')
for row in ws.values:
    for i in range(len(row)):
        row[i]=i
    print('----------------------------')


wb.save('电子表格测试.xlsx')
isinstance(object, type)