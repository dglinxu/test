import openpyxl
from openpyxl.styles import Font,colors,Alignment

# # 打开存在的表格
wb=openpyxl.load_workbook('openpyxl测试表格.xlsx')
#取得工作簿中表的办法
ws=wb.worksheets[1] #索引
# ws=wb.get_sheet_by_name('2020-02-05') #根据sheet名

# ws.title='第一表' #更改表名
# ws.sheet_properties_tabColor="1072BA"

# wb_sn=wb.sheetnames #取得工作簿中所有的表名
# 最大行，最大列
# print(ws.max_row)
# print(ws.max_column)

# 访问单元格的办法
# ws['b1']='B1单元格'
# ws.cell(row=2,column=1,value='A2单元格')
# ws[1][0].value #第一行，第一个单元格
# ws.append(('第四行','B4')) #增加数据列表和元组都行

# ws.insert_rows(2) #在第二行前插入一行
# ws.insert_rows(4,3) #4行前插入3列
# ws.delete_rows(5) #删除一行
# ws.insert_cols(2,3) #第二列前插入3列
# ws.delete_cols(1,3) #删除1列及后面3列
# ws.move_range("A3:C10", rows=-1, cols=8) 
#向上移动一行，向右移动8列，目标移动区域现在存在单元格将会被覆写
ws.copy("A3:C10", rows=-1, cols=8)

# 单元格属性设置
# bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)
# btf=Font(name='微软雅黑',bold=True,size=18,color=colors.BLUE)
# ws['A1'].font = bold_itatic_24_font
# ws['b2'].font=btf
# ws['a2'].alignment=Aligment(horizontal='center', vertical='center')
# # ws[3].alignment=Aligment(horizontal='center', vertical='center')
# ws.row_dimensions[2].height = 40
# ws.column_dimensions['C'].width = 30
# ws.merge_cells('B1:d1') # 合并一行中的几个单元格
# ws.merge_cells('a2:C3') # 合并一个矩形区域中的单元格
# ws.unmerge_cells('A2:c3')

# 数据块的访问，注意单元格数据类型
# data=ws['a1':'b4'] #按行打印
# data=ws['a'] 
# print([cell.value for cell in data]) 
# data=ws['a:b'] #注意只有一对单引号，是按列打印,反过来b:a数据为空
# data=ws[1:5]   #先行到列
# data=ws[1]
# print(list(map(lambda x:x.value,data)))
# for row in data: 
#     for cell in row: 
#         print(cell.value,end=' ')
#     print('\n-------')

# 输出全部数据
# for row in ws.rows: 
#     for cell in row: 
#         print(cell.value)
# for column in ws.columns: 
#     for cell in column: 
#         print(cell.value)

#行列转置
# rows = [
#     ['Number', 'data1', 'data2'],
#     [2, 40, 30],
#     [3, 40, 25],
#     [4, 50, 30],
#     [5, 30, 10],
#     [6, 25, 5],
#     [7, 50, 10]]
# for row in rows: 
#     ws.append(row)
# columns=list(zip(*rows))
# for column in columns: 
#     ws.append(column)



# wb.close() #操作结束后要保存
# 实例化
# wb =openpyxl.Workbook()
# # 激活 worksheet,默认为第一页
# ws = wb.active
# wb.create_sheet('2020-02-05',0)
# wb.remove('2020-02-05')
# del wb['2020-02-05']
# ws['a1']='营销中心'
# wb.close     #只是表示关闭，不保存
wb.save('openpyxl测试表格.xlsx') #操作结束后要保存，close不行
